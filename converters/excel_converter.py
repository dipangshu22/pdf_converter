"""
Excel to PDF Converter
Converts .xlsx / .xls files to PDF using openpyxl + reportlab.
Preserves: cell alignment (horizontal/vertical), font bold/italic/size/color,
           multiple sheets, auto-column widths, styled headers.
"""

import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles.alignment import Alignment

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table,
    TableStyle, HRFlowable, PageBreak
)
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT


SUPPORTED_FORMATS = {'.xlsx', '.xls', '.xlsm'}

MAX_COLS_PORTRAIT = 8
MAX_CELL_WIDTH    = 60

# Map openpyxl horizontal alignment → ReportLab TA_*
_H_ALIGN_MAP = {
    'left':      TA_LEFT,
    'center':    TA_CENTER,
    'right':     TA_RIGHT,
    'general':   TA_LEFT,
    'justify':   TA_LEFT,
    'fill':      TA_LEFT,
    None:        TA_LEFT,
}
_H_ALIGN_RL = {
    TA_LEFT:   'LEFT',
    TA_CENTER: 'CENTER',
    TA_RIGHT:  'RIGHT',
}


def _cell_value(cell) -> str:
    if cell.value is None:
        return ''
    return str(cell.value)


def _cell_rl_align(cell) -> tuple:
    """Return (TA_*, 'LEFT'|'CENTER'|'RIGHT') for a cell."""
    try:
        h = (cell.alignment.horizontal or 'general') if cell.alignment else 'general'
    except Exception:
        h = 'general'
    ta = _H_ALIGN_MAP.get(h, TA_LEFT)
    return ta, _H_ALIGN_RL.get(ta, 'LEFT')


def _cell_font_props(cell) -> dict:
    """Return bold, italic, size, color for a cell's font."""
    props = {'bold': False, 'italic': False, 'size': 8.0, 'color': None}
    try:
        f = cell.font
        if f:
            props['bold']   = bool(f.bold)
            props['italic'] = bool(f.italic)
            props['size']   = float(f.sz or 8)
            if f.color and f.color.type == 'rgb' and f.color.rgb:
                rgb = str(f.color.rgb)
                # Excel ARGB is 8 hex chars; strip alpha
                if len(rgb) == 8:
                    rgb = rgb[2:]
                if len(rgb) == 6:
                    props['color'] = f'#{rgb}'
    except Exception:
        pass
    return props


def _get_col_widths(ws, used_cols: int, page_width_pts: float) -> list:
    col_lengths = []
    for col_idx in range(1, used_cols + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter]:
            val = _cell_value(cell)
            max_len = max(max_len, min(len(val), MAX_CELL_WIDTH))
        col_lengths.append(max(max_len, 4))

    total = sum(col_lengths) or 1
    margin = 40 * mm
    avail  = page_width_pts - margin
    return [avail * (l / total) for l in col_lengths]


def _sheet_to_flowables(ws, sheet_name: str, base_styles) -> list:
    flowables = []

    title_style = ParagraphStyle(
        'SheetTitle',
        parent=base_styles['Heading2'],
        fontSize=14, leading=18,
        textColor=colors.HexColor('#1a1a2e'),
        fontName='Helvetica-Bold',
        spaceBefore=8, spaceAfter=6,
    )
    flowables.append(Paragraph(f'Sheet: {sheet_name}', title_style))
    flowables.append(HRFlowable(width='100%', thickness=1.5, color=colors.HexColor('#2c3e50')))
    flowables.append(Spacer(1, 4))

    min_row = ws.min_row or 1
    max_row = ws.max_row or 1
    min_col = ws.min_column or 1
    max_col = ws.max_column or 1

    if max_row < min_row or max_col < min_col:
        flowables.append(Paragraph('(Empty sheet)', base_styles['Normal']))
        return flowables

    used_cols  = max_col - min_col + 1
    page_size  = landscape(A4) if used_cols > MAX_COLS_PORTRAIT else A4
    page_width = page_size[0]
    col_widths = _get_col_widths(ws, used_cols, page_width)

    # Build table data + per-cell style commands
    table_data    = []
    align_cmds    = []   # ('ALIGN', ...)
    font_cmds     = []   # ('FONTNAME', ...) / ('FONTSIZE', ...) / ('TEXTCOLOR', ...)
    bg_cmds       = []   # ('BACKGROUND', ...) for colored cells

    for row_idx, row in enumerate(ws.iter_rows(
            min_row=min_row, max_row=max_row,
            min_col=min_col, max_col=max_col)):
        row_data = []
        for col_idx, cell in enumerate(row):
            ta_align, rl_align = _cell_rl_align(cell)
            fp = _cell_font_props(cell)

            val = _cell_value(cell)

            # Build ReportLab font name
            fn = 'Helvetica-Bold' if fp['bold'] else 'Helvetica'
            if fp['italic'] and fp['bold']:
                fn = 'Helvetica-BoldOblique'
            elif fp['italic']:
                fn = 'Helvetica-Oblique'

            cell_style = ParagraphStyle(
                'CS',
                fontSize=fp['size'],
                leading=fp['size'] * 1.3,
                wordWrap='CJK',
                alignment=ta_align,
                fontName=fn,
            )
            row_data.append(Paragraph(val, cell_style))

            coord = (col_idx, row_idx)
            align_cmds.append(('ALIGN', coord, coord, rl_align))
            font_cmds.append(('FONTNAME', coord, coord, fn))
            font_cmds.append(('FONTSIZE', coord, coord, fp['size']))
            if fp['color']:
                font_cmds.append(('TEXTCOLOR', coord, coord, colors.HexColor(fp['color'])))

            # Cell background color
            try:
                fill = cell.fill
                if fill and fill.fill_type == 'solid' and fill.fgColor:
                    rgb = str(fill.fgColor.rgb)
                    if len(rgb) == 8:
                        rgb = rgb[2:]
                    if len(rgb) == 6 and rgb.upper() != '000000' and rgb.upper() != 'FFFFFF':
                        bg_cmds.append(('BACKGROUND', coord, coord, colors.HexColor(f'#{rgb}')))
            except Exception:
                pass

        table_data.append(row_data)

    if not table_data:
        flowables.append(Paragraph('(Empty sheet)', base_styles['Normal']))
        return flowables

    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    base_cmds = [
        # Header row
        ('BACKGROUND',    (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR',     (0, 0), (-1, 0), colors.white),
        ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, 0), 9),
        # Alternating rows
        ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, colors.HexColor('#eef2f7')]),
        # Grid
        ('GRID',          (0, 0), (-1, -1), 0.4, colors.HexColor('#c8d0da')),
        # Padding
        ('TOPPADDING',    (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING',   (0, 0), (-1, -1), 5),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 5),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
    ]

    table.setStyle(TableStyle(base_cmds + bg_cmds + align_cmds + font_cmds))
    flowables.append(table)
    return flowables


def convert_excel_to_pdf(input_path: str, output_path: str) -> dict:
    ext = os.path.splitext(input_path)[1].lower()
    if ext not in SUPPORTED_FORMATS:
        return {'success': False, 'message': f"Unsupported format '{ext}'."}

    try:
        wb = openpyxl.load_workbook(input_path, data_only=True)
        sheets = wb.sheetnames
        if not sheets:
            return {'success': False, 'message': 'Excel file has no sheets.'}

        max_cols = max((wb[s].max_column or 0) for s in sheets)
        page_size = landscape(A4) if max_cols > MAX_COLS_PORTRAIT else A4

        doc = SimpleDocTemplate(
            output_path, pagesize=page_size,
            leftMargin=15*mm, rightMargin=15*mm,
            topMargin=15*mm,  bottomMargin=15*mm,
        )

        base_styles = getSampleStyleSheet()
        story = []

        for i, sheet_name in enumerate(sheets):
            ws = wb[sheet_name]
            if ws.max_row is None or ws.max_row == 0:
                continue
            if i > 0:
                story.append(PageBreak())
            story.extend(_sheet_to_flowables(ws, sheet_name, base_styles))

        if not story:
            story.append(Paragraph('(No data found in Excel file)', base_styles['Normal']))

        doc.build(story)
        size_kb = os.path.getsize(output_path) / 1024
        sheet_label = f"{len(sheets)} sheet{'s' if len(sheets) > 1 else ''}"
        return {
            'success': True,
            'message': f"Excel converted successfully ({sheet_label}). Output size: {size_kb:.1f} KB"
        }

    except Exception as e:
        return {'success': False, 'message': f"Excel conversion failed: {str(e)}"}